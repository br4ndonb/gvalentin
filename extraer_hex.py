from openpyxl import load_workbook

def leer_hoja_excel(archivo, nombre_hoja):
    """Lee una hoja específica de un archivo de Excel y devuelve su contenido como matriz."""
    wb = load_workbook(archivo, data_only=True)  # Carga el archivo sin evaluar fórmulas
    
    if nombre_hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo.")

    ws = wb[nombre_hoja]
    matriz = []

    for fila in ws.iter_rows(values_only=True):
        matriz.append(list(fila))  # Convertimos la fila en lista y la agregamos a la matriz

    return {f"particulas_{nombre_hoja}": matriz}  # Usamos el nombre de la hoja en la variable

def exportar_a_txt(datos, archivo_salida):
    """Exporta la matriz de la hoja seleccionada a un archivo de texto."""
    with open(archivo_salida, "w", encoding="utf-8") as f:
        for nombre, matriz in datos.items():
            f.write(f"particulas_ = {matriz};\n\n")  # Formato de salida

# 📌 Uso del código
archivo_excel = "./excel/SANVALENTIN_2.xlsx"  # Reemplázalo con el nombre de tu archivo
archivo_txt = "./colors/colors_variable.txt"
nombre_hoja = "Hoja 1"  # Reemplázalo con la hoja que deseas leer

try:
    datos = leer_hoja_excel(archivo_excel, nombre_hoja)
    exportar_a_txt(datos, archivo_txt)
    print(f"📄 Archivo exportado exitosamente como {archivo_txt}")
except ValueError as e:
    print(f"⚠️ Error: {e}")

